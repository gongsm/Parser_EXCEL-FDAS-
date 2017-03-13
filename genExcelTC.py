'''
Created on 16.10.2015

@author: hud
'''

import re
import time

from openpyxl import Workbook, load_workbook
from openpyxl.cell import get_column_letter, coordinate_from_string, column_index_from_string
from openpyxl.styles import Color, Font, Side, Border, PatternFill, Style, Alignment
from alertExpression import compileToPostfix
from OrderedDict import OrderedDict

#import genDCTable 
import genTruthTable

yellow     = Color(rgb='00ffff00')
red        = Color(rgb='ff0000')
green      = Color(rgb='00b050')
blue       = Color(rgb='0070c0')
white      = Color(rgb='ffffff')
black      = Color(rgb='000000')
gray       = Color(rgb='bfbfbf')
truecolor  = Color(rgb='ebf1de')
falsecolor = Color(rgb='f2dcdb')

thinline = Side(style='thin')
mediumline = Side(style='medium')
thinborder   = Border(left=thinline, right=thinline, top=thinline, bottom=thinline)
mediumborder   = Border(left=mediumline, right=mediumline, top=mediumline, bottom=mediumline)
startmediumborder   = Border(left=thinline, right=thinline, top=mediumline, bottom=thinline)
endmediumborder   = Border(left=thinline, right=thinline, top=thinline, bottom=mediumline)

hdrstyle = Style(
    font=Font(b=True),
    fill=PatternFill(patternType='solid', bgColor=yellow, fgColor=yellow),
    border=thinborder
)

normstyle=Style(border=thinborder)


starttcstyle = Style(font=Font(name='GE Inspira', size=11, b=True, color=red),
                fill=PatternFill(patternType='solid', bgColor=white, fgColor=white),
                border=startmediumborder)

tcstyle=Style(font=Font(name='GE Inspira', size=11),
              alignment=Alignment(horizontal='left', vertical='top', wrap_text=True),
              border=thinborder)

endtcstyle = Style(font=Font(name='GE Inspira', size=11, b=True, color=red),
                fill=PatternFill(patternType='solid', bgColor=white, fgColor=white),
                border=endmediumborder)

startpromptstyle = Style(font=Font(name='GE Inspira', size=11, b=True, color=green),
                fill=PatternFill(patternType='solid', bgColor=white, fgColor=white),
                border=thinborder,
                alignment=Alignment(vertical='top'))
promptstyle=Style(font=Font(name='GE Inspira', size=11),
                 alignment=Alignment(vertical='top', wrap_text=True),
                 border=thinborder)
endpromptstyle = startpromptstyle

scriptstyle = Style(font=Font(name='GE Inspira', size=11, b=True, color=blue),
                fill=PatternFill(patternType='solid', bgColor=white, fgColor=white),
                border=thinborder,
                alignment=Alignment(horizontal='left', vertical='top'))

start_truthtable_style = Style(font=Font(name='GE Inspira', size=12, b=True, color=black),
                            fill=PatternFill(patternType='solid', bgColor=gray, fgColor=gray),
                            border=thinborder,
                            alignment=Alignment(vertical='top'))

centercellstyle = Style(font=Font(name='GE Inspira', size=10),
                      alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                      border=thinborder)

cellstyle = Style(font=Font(name='GE Inspira', size=10),
                      alignment=Alignment(horizontal='left', vertical='top', wrap_text=False),
                      border=thinborder)

cellgraystyle = Style(font=Font(name='GE Inspira', size=10),
                      fill=PatternFill(patternType='solid', bgColor=gray, fgColor=gray),
                      alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                      border=thinborder)

truecellstyle = centercellstyle

truecellmarkstyle = Style(font=Font(name='GE Inspira', size=10),
                          fill=PatternFill(patternType='solid', bgColor=truecolor, fgColor=truecolor),
                          alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                          border=thinborder)

falsecellstyle = centercellstyle

falsecellmarkstyle = Style(font=Font(name='GE Inspira', size=10),
                          fill=PatternFill(patternType='solid', bgColor=falsecolor, fgColor=falsecolor),
                          alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                          border=thinborder)

START_OF_TEST_CASE   = 'START OF TEST CASE'
START_OF_TEST_PROMPT = 'START OF TEST PROMPT'
END_OF_TEST_PROMPT   = 'END OF TEST PROMPT'
START_OF_TEST_SCRIPT = 'START OF TEST SCRIPT'
END_OF_TEST_SCRIPT   = 'END OF TEST SCRIPT'
END_OF_TEST_CASE     = 'END OF TEST CASE'

class Bunch(object):
    def __init__(self, **kwds):
        self.__dict__.update(kwds)
        
def readTrueFalseTable(filename, title='MCDC'):
    try:
        wb = load_workbook(filename, read_only=True, data_only=True)
    except:
        raise Exception, 'Open file failed %' % filename

    ws = wb.get_sheet_by_name(title)
    if not ws:
        raise Exception, 'Sheet MCDC not found'

    logiclist = []
    tflist = []
    plist = []
    originallogic = None
    skip = False
    refid = None
    inhibit = None

    currtf = []
    for row in ws.iter_rows():
        if currtf:
            tflist.append(currtf)
            currtf = []
        if row[0].value is None:
            logiclist.append(Bunch(refid=refid, inhibit=inhibit, logic=originallogic, plist=plist, tflist=tflist, skip=skip))
            tflist = []
            plist = []
            continue
        for cell in row:
            if cell.value is None:
                break
            elif str(cell.value).startswith('#'):
                originallogic = cell.value[1:]
                skip=True
                break
            elif type(cell.value) is bool:
                currtf.append(str(cell.value))
                continue
            elif cell.value is not None and re.match('^[oO][uU][tT]\s*=\s*', cell.value.strip()) is not None:
                originallogic = cell.value
                break
            elif cell.value is not None and re.match('[a-zA-Z]*-[0-9]*([0-9]*)', cell.value.strip()):
                refid = cell.value
                try:
                    inhibit = str(row[1].value)
                except:
                    # flight phase inhibit is None
                    pass
                break
            else:
                plist.append(str(cell.value))

    return logiclist


    # e.g.
    '''
    Bunch(name='010', 
          promptdict=OrderedDict([('TESTING', ('The opFDASAlertBuffer is set to Valid and the value of corresponding bits of opFDASAlertBuffer[17/4] is set to ACTIVE_UNINHIBITED, the opFDASCommandedSynopticPage is set to Valid and AIR, the opFDASMapFeatureRequest is set to Valid and NONE, and the opFDASCollapseAndHoldMenus is set to Valid and FALSE when A1_BLEED_FAULT is set to FALSE, B1_BLEED_FAULT is set to FALSE, A1_LABEL271_Status is set to GLOB_DATA_NO_DATA, B1_LABEL271_Status is set to GLOB_DATA_NO_DATA and the Flight Phase is set to POWER_UP.',)),
                                  ('REQUIREMENTS TESTED', ('FDAS_HLR_47 (COM)', 
                                                           'FDAS_HLR_55 (COM)',
                                                           'FDAS_HLR_284 (COM)',
                                                           'FDAS_HLR_287 (COM)',
                                                           'FDAS_HLR_294 (COM)',
                                                           'FDAS_HLR_291 (COM)',
                                                           'FDAS_HLR_338 (COM)',
                                                           'FDAS_HLR_339 (COM)',
                                                           'FDAS_HLR_344 (COM)',
                                                           'AMS-1 (17)')),
                                  ('ACTION', ()),
                                  ('ENSURE', ()),
                                 ]),
          testscript=(('COMMENT_PYTHON', 'Set the Flight Phase to POWER_UP'),
                      ('CALL_FUNCTION', 'FDAS_FlightPhase', 'SetFlightPhase', 'POWER_UP'),
                      ('COMMENT_PYTHON', 'Set A1_BLEED_FAULT to FALSE, B1_BLEED_FAULT to FALSE, A1_LABEL271_Status to GLOB_DATA_NO_DATA and B1_LABEL271_Status to GLOB_DATA_NO_DATA'),
                      ('SET', 'A1_BLEED_FAULT', 'FALSE'),
                      ('SET', 'B1_BLEED_FAULT', 'FALSE'),
                      ('SET', 'A1_LABEL271_Statue', 'GLOB_DATA_NO_DATA'),
                      ('SET', 'B1_LABEL271_Statue', 'GLOB_DATA_NO_DATA'),
                      ('CALL_FUNCTION', 'FDAS_AlertBuffer', 'CheckAlertBuffer', '17', 'ACTIVE_UNINHIBITED'),
                      ('CHECK', 'opFDASCommandedSynopticPage_Status', 'EQUALS', 'GLOB_DATA_NORMAL_OPERATION'),
                      ('CHECK', 'opFDASCommandedSynopticPage', 'EQUALS', 'AIR'),
                      ('CHECK', 'opFDASMapFeatureRequest_Status', 'EQUALS', 'GLOB_DATA_NORMAL_OPERATION'),
                      ('CHECK', 'opFDASMapFeatureRequest', 'EQUALS', 'NONE'),
                      ('CHECK', 'opFDASCollapseAndHoldMenus_Status', 'EQUALS', 'GLOB_DATA_NORMAL_OPERATION'),
                      ('CHECK', 'opFDASCollapseAndHoldMenus', 'EQUALS', 'FALSE'),),
                     )
    '''

def genTestCaseObject_lv0(tftable, baseidx):
    '''
    testleve: 0 only MCDC; 1 CC + FlightPhase; 2 CC + Umbrella; 3 CC + Collector;
    '''

    res = []
    
    for tf in tftable.tflist:
        name = '%03d' % baseidx
    
        promptdict = OrderedDict()
        promptdict['TESTING'] = ()
        promptdict['REQUIREMENTS TESTED'] = (tftable.refid,)
        promptdict['ACTION'] = ()
        promptdict['ENSURE'] = ()
        
        testscript = []

        teststep = ['CALL_FUNCTION', 'FDAS_FlightPhase', 'SetFlightPhase', 'POWER_UP']
        testscript.append(teststep)

        idx = 0
        for b in tf[:-1]:
            pname = tftable.plist[idx]
            if pname.startswith('.invalid'): # .invalid(AMS.A1_LABEL271)
                pname = pname.split('.')[-1][0:-1] + '_Status'
                if b == 'True':
                    testscript.append(['SET', pname, 'GLOB_DATA_NO_DATA'])
                else:
                    testscript.append(['SET', pname, 'GLOB_DATA_NORMAL_OPERATION'])
            else: # AMS.A1_LABEL271
                pname = pname.split('.')[-1]
                testscript.append(['SET', pname, b])
            idx += 1
            
        teststep = ['CHECK', 'opFDASAlertBuffer_Status', 'EQUALS', 'GLOB_DATA_NORMAL_OPERATION']
        testscript.append(teststep)

        alertid = tftable.refid.split(' ')[1][1:-1]
        if tf[-1] == 'True':
            teststep = ['CALL_FUNCTION', 'FDAS_AlertBuffer', 'CheckAlertBuffer', '%s'%alertid, 'ACTIVE_UNINHIBITED']
        else:
            teststep = ['CALL_FUNCTION', 'FDAS_AlertBuffer', 'CheckAlertBuffer', '%s'%alertid, 'INACTIVE']
        testscript.append(teststep)

        res.append(Bunch(name=name, promptdict=promptdict, testscript=testscript))
        baseidx += 1
        
    return baseidx, res

def genTestCaseObject_lv1(tftable, baseidx):
    '''
    testleve: 0 only MCDC; 1 CC + FlightPhase; 2 CC + Umbrella; 3 CC + Collector;
    '''
    
    res = []
    flightphase = ['POWER_UP', 'TAXI_OUT', 'TO1', 'TO2', 'TO3', 'CRUISE', 'APPROACH', 'LANDING', 'TAXI_IN', 'SHUT_DOWN',] 
    
    # Selet one True and one False condition
    tcondition = None
    fcondition = None
    for tf in tftable.tflist:
        if tf[-1] == 'True':
            # True condition
            tcondition = tf
        else:
            # False condition
            fcondition = tf
        if tcondition is not None and fcondition is not None:
            break
        
    if tcondition is None or fcondition is None:
        raise Exception, '%s' % tf

    for fp in flightphase:
        name = '%03d' % baseidx
    
        promptdict = OrderedDict()
        promptdict['TESTING'] = ()
        promptdict['REQUIREMENTS TESTED'] = (tftable.refid,)
        promptdict['ACTION'] = ()
        promptdict['ENSURE'] = ()
        
        testscript = []

        teststep = ['CALL_FUNCTION', 'FDAS_FlightPhase', 'SetFlightPhase', '%s'%fp]
        testscript.append(teststep)

        # True condition
        idx = 0
        for b in tcondition[:-1]:
            pname = tftable.plist[idx]
            if pname.startswith('.invalid'): # .invalid(AMS.A1_LABEL271)
                pname = pname.split('.')[-1][0:-1] + '_Status'
                if b == 'True':
                    testscript.append(['SET', pname, 'GLOB_DATA_NO_DATA'])
                else:
                    testscript.append(['SET', pname, 'GLOB_DATA_NORMAL_OPERATION'])
            else: # AMS.A1_LABEL271
                pname = pname.split('.')[-1]
                testscript.append(['SET', pname, b])
            idx += 1
        
        teststep = ['CHECK', 'opFDASAlertBuffer_Status', 'EQUALS', 'GLOB_DATA_NORMAL_OPERATION']
        testscript.append(teststep)

        alertid = tftable.refid.split(' ')[1][1:-1]
        inhibit = tftable.inhibit.upper().split(' ')
        if fp in inhibit:
            # ACTIVE_INHIBITED
            teststep = ['CALL_FUNCTION', 'FDAS_AlertBuffer', 'CheckAlertBuffer', '%s'%alertid, 'ACTIVE_INHIBITED']
        else:
            #ACTIVE_UNINHIBITED
            teststep = ['CALL_FUNCTION', 'FDAS_AlertBuffer', 'CheckAlertBuffer', '%s'%alertid, 'ACTIVE_UNINHIBITED']
        testscript.append(teststep)

        # False condition
        idx = 0
        for b in fcondition[:-1]:
            pname = tftable.plist[idx]
            if pname.startswith('.invalid'): # .invalid(AMS.A1_LABEL271)
                pname = pname.split('.')[-1][0:-1] + '_Status'
                if b == 'True':
                    testscript.append(['SET', pname, 'GLOB_DATA_NO_DATA'])
                else:
                    testscript.append(['SET', pname, 'GLOB_DATA_NORMAL_OPERATION'])
            else: # AMS.A1_LABEL271
                pname = pname.split('.')[-1]
                testscript.append(['SET', pname, b])
            idx += 1
        
        teststep = ['CHECK', 'opFDASAlertBuffer_Status', 'EQUALS', 'GLOB_DATA_NORMAL_OPERATION']
        testscript.append(teststep)

        alertid = tftable.refid.split(' ')[1][1:-1]
        # Alwayse INACITVE
        teststep = ['CALL_FUNCTION', 'FDAS_AlertBuffer', 'CheckAlertBuffer', '%s'%alertid, 'INACTIVE']
        testscript.append(teststep)

        res.append(Bunch(name=name, promptdict=promptdict, testscript=testscript))
        baseidx += 1

    return baseidx, res

def genConfigurationSheet(ws):
    pass

def remarkConditions():
    pass

def genTruthTableRow(ws, rowidx, row):
    for cellidx in range(0, len(row)):
        currcell = ws.cell(row=rowidx, column=cellidx+3)
        if row[cellidx] == 'true':
            currcell.value = 'T'
            currcell.style = truecellstyle
        elif row[cellidx] == 'TRUE':
            currcell.value = 'T'
            currcell.style = truecellmarkstyle
        elif row[cellidx] == 'false':
            currcell.value = 'F'
            currcell.style = falsecellstyle
        elif row[cellidx] == 'FALSE':
            currcell.value = 'F'
            currcell.style = falsecellmarkstyle

def genTruthTableCell(ws, rowidx, table):
    startcell1 = ws.cell(row=rowidx, column=1)
    startcell1.value = "REQUIREMENT ID"
    startcell1.style = start_truthtable_style

    startcell2 = ws.cell(row=rowidx, column=2)
    startcell2.value = table.sourceid + " (" + table.alertid + ")"
    startcell2.style = start_truthtable_style
    ws.merge_cells(start_row=rowidx, start_column=2, end_row=rowidx, end_column=len(table.conditions)+2)
    
    rowidx += 1
    
    reqlogiccell1 = ws.cell(row=rowidx, column=1)
    reqlogiccell1.value = "REQUIREMENT LOGIC:"
    reqlogiccell1.style = cellstyle

    reqlogiccell2 = ws.cell(row=rowidx, column=2)
    reqlogiccell2.value = table.logic
    reqlogiccell2.style = cellstyle 
    ws.merge_cells(start_row=rowidx, start_column=2, end_row=rowidx, end_column=len(table.conditions)+2)
    
    rowidx += 1
    
    boollogiccell1 = ws.cell(row=rowidx, column=1)
    boollogiccell1.value = "BOOLEAN LOGIC:"
    boollogiccell1.style = cellstyle

    boollogiccell2 = ws.cell(row=rowidx, column=2)
    simplelogic = table.booleanlogic
    i = 0
    for c in table.conditions:
        simplelogic.replace(c, chr(65+i))
        i += 1
    boollogiccell2.value = simplelogic
    boollogiccell2.style = cellstyle
    ws.merge_cells(start_row=rowidx, start_column=2, end_row=rowidx, end_column=len(table.conditions)+2)
    
    rowidx += 1
    
    tccell = ws.cell(row=rowidx, column=1)
    tccell.value = "TEST CASE"
    tccell.style = cellgraystyle

    tccell = ws.cell(row=rowidx, column=2)
    tccell.value = "OUT"
    tccell.style = cellgraystyle
    
    for cidx in range(0, len(table.conditions)):
        tccell = ws.cell(row=rowidx, column=cidx+3)
        tccell.value = table.conditions[cidx]
        tccell.style = cellgraystyle
    
    rowidx += 1

    for cidx in range(0, len(table.alphabet)):
        tccell = ws.cell(row=rowidx, column=cidx+3)
        tccell.value = table.alphabet[cidx] #chr(cidx+65)
        tccell.style = cellgraystyle
    ws.merge_cells(start_row=rowidx-1, start_column=1, end_row=rowidx, end_column=1)
    ws.merge_cells(start_row=rowidx-1, start_column=2, end_row=rowidx, end_column=2)
    rowidx += 1

    # Cross T/F test cases, like T/F/T/F/T/F...
    tcnum = 1
    tidx = -1
    for tidx in range(0, len(table.truthtable)):
        if table.truthtable[tidx] == []:
            continue
        
        tcnumcell = ws.cell(row=rowidx, column=1)
        tcnumcell.value = '%03d' % (tcnum * 10)
        tcnumcell.style = cellstyle
        
        outcell = ws.cell(row=rowidx, column=2)
        outcell.value = 'T'
        outcell.style = cellstyle
        
        genTruthTableRow(ws, rowidx, table.truthtable[tidx])
        rowidx += 1
        
        if tidx < len(table.falsetable) and table.falsetable[tidx] != []:
            tcnum += 1
            tcnumcell = ws.cell(row=rowidx, column=1)
            tcnumcell.value = '%03d' % (tcnum * 10)
            tcnumcell.style = cellstyle

            outcell = ws.cell(row=rowidx, column=2)
            outcell.value = 'F'
            outcell.style = cellstyle

            genTruthTableRow(ws, rowidx, table.falsetable[tidx])
            rowidx += 1

        tcnum += 1
        
    for idx in range(tidx+1, len(table.falsetable)):
        if table.falsetable[idx] == []:
            continue
        
        tcnumcell = ws.cell(row=rowidx, column=1)
        tcnumcell.value = '%03d' % (tcnum * 10)
        tcnumcell.style = cellstyle
        
        outcell = ws.cell(row=rowidx, column=2)
        outcell.value = 'F'
        outcell.style = cellstyle
        
        genTruthTableRow(ws, rowidx, table.falsetable[idx])
        rowidx += 1
        tcnum += 1
        
    return rowidx

def genTruthTableSheet(ws, data):
    startrowidx = 2
    if data is None:
        return
    currrowidx = startrowidx
    for t in data:
        currrowidx = genTruthTableCell(ws, currrowidx+1, t)

def genTestCase(ws, rowidx, tc):
    # START OF TEST CASE
    starttccell = ws.cell(row=rowidx, column=1)
    starttccell.value = (START_OF_TEST_CASE)
    starttccell.style = starttcstyle
    
    tcnamecell = ws.cell(row=rowidx, column=2)
    tcnamecell.value = (tc.name)
    tcnamecell.style = tcstyle

    rowidx += 1
    blankCell(ws, rowidx, 1)

    # START OF TEST PROMPT
    startpromptcell = ws.cell(row=rowidx, column=1)
    startpromptcell.value = (START_OF_TEST_PROMPT)
    startpromptcell.style = startpromptstyle
    
    for promptkey, prompts in tc.promptdict.iteritems():
        promptcell = ws.cell(row=rowidx, column=2)
        promptcell.value = promptkey
        promptcell.style = promptstyle
        
        for promptstr in prompts:
            promptcell = ws.cell(row=rowidx, column=3)
            promptcell.value = promptstr
            promptcell.style = promptstyle
            rowidx += 1
            blankCell(ws, rowidx, 1)
            
        if not prompts:
            rowidx += 1
            blankCell(ws, rowidx, 1)

        rowidx += 1
        blankCell(ws, rowidx, 1)
    
    rowidx += 1
    blankCell(ws, rowidx, 1)

    # END OF TEST PROMPT
    endpromptcell = ws.cell(row=rowidx, column=1)
    endpromptcell.value = (END_OF_TEST_PROMPT)
    endpromptcell.style = endpromptstyle
    rowidx += 2
    blankCell(ws, rowidx, 1)

    # START OF TEST SCRIPT
    startscriptcell = ws.cell(row=rowidx, column=1)
    startscriptcell.value = (START_OF_TEST_SCRIPT)
    startscriptcell.style = scriptstyle
    rowidx += 1
    blankCell(ws, rowidx, 1)
    
    stepidx = 1
    #print tc.testscript
    for teststep in tc.testscript:
        stepidxcell = ws.cell(row=rowidx, column=1)
        stepidxcell.value = stepidx
        stepidxcell.style = scriptstyle
        
        stepcolumn = 2
        for stepstr in teststep:
            #print stepstr
            stepcell = ws.cell(row=rowidx, column=stepcolumn)
            stepcell.value = stepstr
            stepcell.style = tcstyle
            stepcolumn += 1
        
        stepidx += 1
        rowidx += 1
        blankCell(ws, rowidx, 1)

    # END OF TEST SCRIPT
    endscriptcell = ws.cell(row=rowidx, column=1)
    endscriptcell.value = (END_OF_TEST_SCRIPT)
    endscriptcell.style = scriptstyle
    rowidx += 1
    blankCell(ws, rowidx, 1)

    # END OF TEST CASE
    endtccell = ws.cell(row=rowidx, column=1)
    endtccell.value = (END_OF_TEST_CASE)
    endtccell.style = endtcstyle
    rowidx += 1
    
    return rowidx

def genSoftwareSheet(ws):
    pass

def blankCell(ws, row, column):
    blankcell = ws.cell(row=row, column=column)
    blankcell.style = normstyle
    
def genTestCaseSheet(ws, tcname, headdesc, tclist):
    idx = 1
    for row in ws.iter_rows():
        if str(row[0].value) == 'SCRIPT NAME':
            row[1].value = tcname
        if str(row[0].value) == 'HEADER':
            ws.cell(row=idx+1, column=2).value = headdesc
        if str(row[0].value) == 'END OF HEADER':
            endheadline = idx+1
            break
        idx += 1
    
    # INITIALISATION
    endinitial = genTestCase(ws, endheadline+1, Bunch(name='INITIALISATION', 
                                                          promptdict=OrderedDict([('', ('This is the initialisation test case',
                                                                          'It will be run at the beginning of the test to initialize corresponding parameters to normal status.')),]),
                                                          testscript=(('CALL_FUNCTION', 'FDAS_Init', 'Init'),),
                                                          ))

    applyBorderStyleTestCaseSheet(ws, endheadline+1, endinitial)

    # TC 
    starttcrow = endinitial+1
    #print time.strftime('Step 2-1: %H:%M:%S')
    idx = 0
    for tc in tclist:
        idx += 1
        #if idx % 10 == 0:
        #    print time.strftime('Step 2-1-x: %H:%M:%S')
        endtc = genTestCase(ws, starttcrow, tc)
        applyBorderStyleTestCaseSheet(ws, starttcrow, endtc)
        starttcrow = endtc + 1
    #print time.strftime('Step 2-2: %H:%M:%S')

def border_range(ws, cell_range, style=None):
    start_cell, end_cell = cell_range.split(':')
    start_coord = coordinate_from_string(start_cell)
    start_row = start_coord[1]
    start_col = column_index_from_string(start_coord[0])
    end_coord = coordinate_from_string(end_cell)
    end_row = end_coord[1]
    end_col = column_index_from_string(end_coord[0])

    startcell = ws.cell('%s%s' % (get_column_letter(start_col), start_row))
    if style is None:
        oldstyle = startcell.style
        oldfont = oldstyle.font
        oldfill = oldstyle.fill
        oldborder = oldstyle.border
        oldalignment = oldstyle.alignment
        oldnumber_format = oldstyle.number_format
        oldprotection = oldstyle.protection
    else:
        oldstyle = style
        oldfont = oldstyle.font
        oldfill = oldstyle.fill
        oldborder = startcell.style.border
        oldalignment = oldstyle.alignment
        oldnumber_format = oldstyle.number_format
        oldprotection = oldstyle.protection

    for row in range(start_row, end_row + 1):
        for col_idx in range(start_col, end_col + 1):
            if style is not None and col_idx==start_col:
                continue
            col = get_column_letter(col_idx)
            
            newstyle = Style(font=oldfont, fill=oldfill, border=oldborder, alignment=oldalignment, number_format=oldnumber_format, protection=oldprotection)
            if col_idx == end_col:
                newborder = Border(left=oldborder.left, right=oldborder.left, bottom=oldborder.bottom, diagonal=oldborder.diagonal, diagonal_direction=oldborder.diagonal_direction, vertical=oldborder.vertical, horizontal=oldborder.horizontal)
                newstyle = Style(font=oldfont, fill=oldfill, border=newborder, alignment=oldalignment, number_format=oldnumber_format, protection=oldprotection)

            ws.cell('%s%s' % (col, row)).style = newstyle

def applyMergedCellBorderStyle(ws):
    for cellrange in ws.merged_cell_ranges:
        border_range(ws, cellrange)

def applyBorderStyleTestCaseSheet(ws, starttc, endtc):
    for rowidx in range(starttc, endtc):
        border_range(ws, 'A%s:I%s' % (rowidx, rowidx), style=tcstyle)

def genExcelTC(templatename, ofilename, tcobjlist, truthtable):
    if not templatename.endswith('.xlsx'):
        templatename += '.xlsx'
    try:
        wb = load_workbook(templatename, use_iterators=False)
    except:
        raise Exception, "File open failed %s" % templatename
    
    configurationSheet = wb.worksheets[0]
    truthTableSheet    = wb.worksheets[1]
    testCaseSheet      = wb.worksheets[2]
    explanationSheet   = wb.worksheets[3]

    #genConfigurationSheet(configurationSheet)
    #applyMergedCellBorderStyle(configurationSheet)

    genTruthTableSheet(truthTableSheet, truthtable)
    applyMergedCellBorderStyle(truthTableSheet)

    genTestCaseSheet(testCaseSheet, 'FDAS_AMSAlertDefinition_HLR', "Test Member system AMS's Alert", tcobjlist)
    applyMergedCellBorderStyle(testCaseSheet)

    if not ofilename.endswith('.xlsx'):
        ofilename += '.xlsx'
    wb.save(ofilename)
    
def findOpposite(idx, table, tablelist, coupledconditions):
    for t in tablelist:
        t_tmp = map(lambda n: n.upper(), t[:-1])
        table_tmp = map(lambda n: n.upper(), table[:-1])
        t_tmp[idx] = table_tmp[idx]
        for k, idxlist in coupledconditions.iteritems():
            if idx in idxlist:
                for i in idxlist:
                    t_tmp[i] = table_tmp[i]
        if t_tmp == table_tmp:
            return True
    return False

def convertCondition(idx, table, coupledconditions):
    table[idx] = not table[idx]

    if coupledconditions:
        for idxlist in coupledconditions.values():
            if idx in idxlist:
                for coupledidx in idxlist:
                    if idx == coupledidx:
                        continue
                    table[coupledidx] = not table[coupledidx]

def markMCDCPair(alerts):
    for alert in alerts:
        fmarkdict = {}
        tableidx = 0
        for ft in alert.falsetable:
            cidx = 0
            fmarkdict[tableidx] = []
            for c in ft:
                ft_tmp = map(lambda n: n, ft)
                #ft_tmp[cidx] = not ft_tmp[cidx]
                convertCondition(cidx, ft_tmp, alert.coupledconditions)
                if ft_tmp in alert.truthtable:
                    fmarkdict[tableidx].append(cidx)
                cidx += 1
            tableidx += 1 

        tmarkdict = {}
        tableidx = 0
        for tt in alert.truthtable:
            cidx = 0
            tmarkdict[tableidx] = []
            for c in tt:
                tt_tmp = map(lambda n: n, tt)
                #tt_tmp[cidx] = not tt_tmp[cidx]
                convertCondition(cidx, tt_tmp, alert.coupledconditions)
                if tt_tmp in alert.falsetable:
                    tmarkdict[tableidx].append(cidx)
                cidx += 1
            tableidx += 1
            
        for k, v in fmarkdict.iteritems():
            marked = False
            for idx in range(0, len(alert.falsetable[k])):
                if idx in v:
                    marked = True
                    alert.falsetable[k][idx] = str(alert.falsetable[k][idx]).upper()
                else:
                    alert.falsetable[k][idx] = str(alert.falsetable[k][idx]).lower()
            #if not marked:
            #    alert.falsetable[k] = []

        for k, v in tmarkdict.iteritems():
            marked = False
            for idx in range(0, len(alert.truthtable[k])):
                if idx in v:
                    marked = True
                    alert.truthtable[k][idx] = str(alert.truthtable[k][idx]).upper()
                else:
                    alert.truthtable[k][idx] = str(alert.truthtable[k][idx]).lower()
            #if not marked:
            #    alert.truthtable[k] = []
    
def reworkTruthFalseTable(alerts):
    for alert in alerts:
        for ft in alert.falsetable:
            idx = 0
            for f in ft:
                if f == "FALSE" and not findOpposite(idx, ft, alert.truthtable, alert.coupledconditions):
                    #alerts[alerts.index(alert)].falsetable[alert.falsetable.index(ft)][idx] = "false"
                    ft[idx] = "false"
                idx += 1
                    
        for tt in alert.truthtable:
            idx = 0
            for t in tt:
                if t == "TRUE" and not findOpposite(idx, tt, alert.falsetable, alert.coupledconditions):
                    #alerts[alerts.index(alert)].truthtable[alert.truthtable.index(tt)][idx] = "true"
                    tt[idx] = 'true'
                idx += 1

    return alerts

def usage():
    print "Usage of genExcelTC.py:"
    print "Input: generated Ture/False Table in Excel format"
    print "Output: Excel Test Case"
    
def main():
    tablelist = readTrueFalseTable("../test/FDASLogicTable.xlsx")
    tcobjlist = []
    tcidx = 1
    for table in tablelist:
        tcidx, tcobj = genTestCaseObject_lv0(table, tcidx)
        tcobjlist.extend(tcobj)
        tcidx += 1

    input_filename = "../test/FDASAlert.xml"
    #(sheet, alerts) = genDCTable.genTruthTable(input_filename)
    alerts = genTruthTable.main(input_filename)
    
    #alerts = reworkTruthFalseTable(alerts)
    markMCDCPair(alerts)
    genExcelTC("../test/Test_Case_Template", "../test/FDAS_AMSAlertDefinition_HLR", tcobjlist, alerts)

if __name__ == "__main__":
    import cProfile
    print time.strftime('Step 1: %H:%M:%S')
    cProfile.run('main()', 'stat.data')
    #main()
    print time.strftime('Step 2: %H:%M:%S')
    
    #TESTING
    #print time.strftime('%x %X %Z')
    #print time.strftime('Step 3: %H:%M:%S')
    #print time.strftime('%x %X %Z')

    #tablelist = readTrueFalseTable("../test/FDASLogicTable.xlsx")
    #tcobjlist = []
    #tcidx = 1
    #for table in tablelist:
    #    tcidx, tcobj = genTestCaseObject_lv1(table, tcidx)
    #    tcobjlist.extend(tcobj)
    #    tcidx += 1
    #genExcelTC("../test/FDAS_AlertDefinition_HLR_Template", "../test/FDAS_AMSAlertDefinition_HLR_FlightPhase", tcobjlist)
        
